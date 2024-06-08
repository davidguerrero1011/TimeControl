import datetime

import bcrypt
import openpyxl
import pandas as pd
from django.contrib.auth.hashers import make_password, check_password
from django.core import serializers
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from pandas.core.interchange import column
from pyexpat.errors import messages
from tablib import Dataset

from employeeapp.models import Employees, EventsType, TimesControl, TypeDocuments, \
    Genre, Roles, TimesControlEvents

from django.contrib import messages

from employeeapp.resource import TimesControlResource
from openpyxl.reader.excel import load_workbook


def welcome(request):
    title = {'title': 'Login'}
    return render(request, 'login/login.html', {'title': title})



def validateCredentials(request):
    if request.method == 'POST':
        users = Employees.objects.filter(username=request.POST['usuario'], password=request.POST['contrasena'])
        countUsers = users.count()

        if countUsers > 0:
            request.session['id'] = users[0].id
            request.session['name'] = users[0].name + ' ' + users[0].last_name
            request.session['identifier'] = users[0].identifier
            request.session['username'] = users[0].username
            request.session['role'] = users[0].role_id

            messages.success(request, f'{request.session["name"]} ingreso con exito, bienvenido una vez más')
            return redirect('index')
        else:
            messages.error(request, "Algo sucedio, pero no puede ingresar, por favor revise sus credenciales de ingreso")
            return redirect('/')



def logoutSession(request):
    request.session['id'] = ''
    request.session['name'] = ''
    request.session['identifier'] = ''
    request.session['username'] = ''
    request.session['role'] = ''

    messages.success(request, f'¡Muchas gracias por haber ingresado en nuestro sitio web, te esperamos nuevamente por aca!')
    return  redirect('/')



def index(request):
    title = 'Bienvenido '
    users = Employees.objects.select_related('genre', 'role', 'type_document').order_by('id')
    events = EventsType.objects.all()
    return render(request, 'layout/index.html', { 'title': title, 'users': users, 'events': events })



def show(request, id):
    if not id or id == 0:
        render(request, 'resources/error/error-404.html')
    else:
        user = Employees.objects.select_related('genre', 'role', 'type_document').get(pk=id)
        return render(request, 'layout/resources/show.html', {'user': user})



def create(request):
    type = 1
    title = 'Crear Usuario'
    type_documents = TypeDocuments.objects.filter(is_active=True)
    genres = Genre.objects.filter(is_active=True)
    roles = Roles.objects.filter(is_active=True)
    return render(request, 'layout/resources/resource.html', {'type': type, 'type_documents': type_documents, 'genres': genres, 'roles': roles, 'title': title})



def edit(request, id):
    type = 2
    title = 'Editar'
    user = Employees.objects.get(pk=id)
    type_documents = TypeDocuments.objects.filter(is_active=True)
    genres = Genre.objects.filter(is_active=True)
    roles = Roles.objects.filter(is_active=True)
    return render(request, 'layout/resources/resource.html', {'type': type, 'title': title, 'type_documents': type_documents, 'genres': genres, 'roles': roles, 'user': user})



def update(request):
    try:
        update_user = Employees.objects.filter(id=request.POST['userId']).update(
                                                                                 name=request.POST['userName'],
                                                                                 last_name=request.POST['lastNames'],
                                                                                 identifier=request.POST['numbIdentifi'],
                                                                                 email=request.POST['userEmail'],
                                                                                 username=request.POST['nameIdentification'],
                                                                                 is_active='true' if request.POST['statusUser'] == 1 else 'false'   ,
                                                                                 updated_at=datetime.datetime.now().today(),
                                                                                 genre_id=request.POST['userGenre'],
                                                                                 role_id=request.POST['userRole'],
                                                                                 type_document_id=request.POST['typeDocu']
                                                                                 )
        messages.success(request, 'El usuario fue actualizado con exito.')
        return redirect(request.META.get('HTTP_REFERER'))

    except FileNotFoundError:
        messages.error(request, 'Hubo algún error actualizando los datos del usuario, deberia arreglarlo')



def changePassword(request):
    try:
        hashed_pwd = make_password(request.POST['passwordNew'])
        Employees.objects.filter(id=request.POST['userId']).update(password=hashed_pwd)

        messages.success(request, f'La contraseña fue actualizada con exito')
        return  redirect(request.META.get('HTTP_REFERER'))
    except FileNotFoundError:
        messages.error('Hubo algún error actualizando la contraseña, por favor revise')



def showTimeRegisters(request, id):
    title = 'Registros de '
    events = EventsType.objects.filter(is_active=True).values('id', 'name')
    registers_by_employee  = TimesControlEvents.objects.select_related('user', 'event', 'time_control').filter(user_id = id)
    return render(request, 'layout/resources/registers.html', {'title': title, 'registers': registers_by_employee, 'events': events})



def showRegister(request):
    single_register = TimesControlEvents.objects.select_related('user', 'event', 'time_control').filter(pk=request.id)
    return HttpResponse(single_register)



def storeTimeUser(request):

    eventRequest = request.POST['event']
    identificationRequest = request.POST['identification']
    if request.method == 'POST':
        if eventRequest is not None or identificationRequest is not None:

            user = Employees.objects.filter(identifier=request.POST['identification'])
            if request.POST['event'] == '1' or request.POST['event'] == '2' or request.POST['event'] == '4' or request.POST['event'] == '5':

                timeControl = TimesControl()
                timeControl.start_date = datetime.datetime.now().time()
                timeControl.save()

                time_id = TimesControl.objects.latest('id')

                event_time_control                 = TimesControlEvents()
                event_time_control.created_at      = datetime.datetime.now().time()
                event_time_control.updated_at      = datetime.datetime.now().time()
                event_time_control.event_id        = request.POST['event']
                event_time_control.time_control_id = time_id.id
                event_time_control.user_id         = user[0].id
                event_time_control.save()

            elif request.POST['event'] == '7' or request.POST['event'] == '8' or request.POST['event'] == '9' or request.POST['event'] == '10':

                timeControl = TimesControl()
                timeControl.endDate = datetime.datetime.now().time()
                timeControl.save()

                time_id = TimesControl.objects.latest('id')

                event_time_control = TimesControlEvents()
                event_time_control.created_at = datetime.datetime.now().time()
                event_time_control.updated_at = datetime.datetime.now().time()
                event_time_control.event_id = request.POST['event']
                event_time_control.time_control_id = time_id.id
                event_time_control.user_id = user[0].id
                event_time_control.save()

            else:

                timeControl = TimesControl()
                timeControl.startDate = datetime.datetime.now().time()
                timeControl.endDate = datetime.datetime.now().time()
                timeControl.save()

                time_id = TimesControl.objects.latest('id')

                event_time_control = TimesControlEvents()
                event_time_control.created_at = datetime.datetime.now().time()
                event_time_control.updated_at = datetime.datetime.now().time()
                event_time_control.event_id = request.POST['event']
                event_time_control.time_control_id = time_id.id
                event_time_control.user_id = user[0].id
                event_time_control.save()

    messages.success(request, f'El registro del usuario fue guardado con exito')
    return redirect("index")



def storeUsers(request):
    if request.method == 'POST':
        salt = bcrypt.gensalt()
        password = request.POST['passwordd']

        employee = Employees()
        employee.name = request.POST['names']
        employee.last_name = request.POST['lastNames']
        employee.identifier = request.POST['documentNumbers']
        employee.email = request.POST['email']
        #employee.password = bcrypt.hashpw(password.encode('utf-8'), salt)
        employee.password = request.POST['passwordd']
        employee.username = request.POST['username']
        employee.is_active = request.POST['status']
        employee.genre_id = request.POST['genre']
        employee.role_id = request.POST['role']
        employee.type_document_id = request.POST['typeDocuments']
        employee.save()

        return redirect('index')



def processFileRegisters(request) :

    #varifico si la peticion viene por post y ademas si viene el archivo con el nombre del campo
    try:
        if request.method == 'POST' and request.FILES['processFile']:

            wb = openpyxl.load_workbook(request.FILES['processFile'])
            ws = wb.active
            data = []
            data.append([ws.cell(row=2,column=i).value for i in range(1,ws.max_column+1)])
            data.append([ws.cell(row=3,column=i).value for i in range(1,ws.max_column+1)])

            time_control_table = TimesControl()
            for i in data:
                time_control_table.start_date = i[2]
                time_control_table.end_date = None if i[3] is None else i[3]
                time_control_table.created_at = datetime.datetime.now().date()
                time_control_table.updated_at = datetime.datetime.now().date()
                time_control_table.save()

            time_id = TimesControl.objects.latest('id')
            time_events_control = TimesControlEvents()
            for x in data:

                userId = Employees.objects.filter(identifier=x[0])

                time_events_control.created_at = datetime.datetime.now().date()
                time_events_control.updated_at = datetime.datetime.now().date()
                time_events_control.event_id = x[1]
                time_events_control.time_control_id = time_id.id
                time_events_control.user_id = userId[0].id
                time_events_control.save()

                messages.success(request, "Eventos almacenados de forma exitosa")
                return redirect('index')
    except FileNotFoundError:
        messages.error(request, "Hubo un error almacenando los registros de los eventos")



def updateTimeRegister(request, id):
    title = f'Actualizacion Individual Registro'
    user_register = TimesControlEvents.objects.select_related('time_control', 'event', 'user').filter(id=id)
    return render(request, 'layout/resources/showRegister.html', {'user_register': user_register, 'title': title})



def deleteRegisters(request, id):
    employeeDelete = Employees.objects.get(pk=id)
    employeeDelete.delete()

    return redirect('index')


